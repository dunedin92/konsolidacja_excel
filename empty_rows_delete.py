import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

def empty_rows_delete(bom_path):
    wb = openpyxl.load_workbook(bom_path)
    type(wb)
    arkusze = wb.sheetnames
    sheet = wb[arkusze[0]]
    max_row = sheet.max_row
    print(max_row)
    sheet.unmerge_cells("A1:Q1")
    sheet.delete_rows(1, 1)



    for i in range(max_row, 1, -1):
        value = sheet.cell(row=i, column=2).value
        print(i, ". - ", value)
        if value != None:
            value = value.lstrip()
        print(value)

        if value == "" or value == None:
            sheet.delete_rows(i, 1)


    print("==> Tworzenie tabeli z całego zakresu danych arkusza.")
    table = Table(displayName="Table1", ref="A1:" + get_column_letter(sheet.max_column) + str(sheet.max_row))
    sheet.add_table(table)
    style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False,
                           showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    table.tableStyleInfo = style

    print("Wstawianie '-' w kazdą pustą komórke w zakresie.")
    for i in range(1, sheet.max_column):
        for j in range(1, sheet.max_row):
            cell_value = sheet.cell(row=j, column=i).value
            print(cell_value)
            if cell_value == "" or cell_value == None:
                sheet.cell(row=j, column=i).value = "-"


    wb.save(bom_path)
    wb.close()