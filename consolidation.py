import openpyxl


class Part:
    def __init__(self, part_number, qty_total=0, description="", description2="", tch="", producent="",
                 kod_producenta="", kolor=""):
        self.part_number = part_number
        self.qty_total = qty_total
        self.description = description
        self.description2 = description2
        self.tch = tch
        self.producent = producent
        self.kod_producenta = kod_producenta
        self.kolor = kolor

    def add_to_qty_total(self, value):
        self.qty_total = self.qty_total + value

    def print_values(self):
        print("part_num = ", self.part_number, "|| qty_tot= ", self.qty_total, "|| desc= ",
              self.description, "|| desc2= ", self.description2, "|| tch= ", self.tch,
              "|| producent= ", self.producent, "|| kod_produc= ", self.kod_producenta, "|| kolor= ", self.kolor)

    def values_to_list(self):
        object_value_list = [self.part_number, self.qty_total, self.description, self.description2,
                             self.producent, self.kod_producenta, self.kolor]
        return object_value_list


def worksheet_maker(bom_path, sheet_name, titles):

    wb = openpyxl.load_workbook(bom_path)
    type(wb)

    status = False
    while not status:
        try:
            wb.save(bom_path)
            status = True
        except PermissionError:
            print("plik jest juz otwarty, zamknij go i kliknij dowolny przycisk")
            input("Naciśnij dowolny klawisz aby kontynuować.")

    arkusze = wb.sheetnames
    if sheet_name in arkusze:
        sheet = wb.get_sheet_by_name(sheet_name)
        wb.remove(sheet)
        sheet = wb.create_sheet(sheet_name)
    else:
        sheet = wb.create_sheet(sheet_name)

    print(sheet)
    column_number = 1
    for title in titles:
        print(title)
        sheet.cell(row=1, column=column_number).value = title
        column_number += 1

    wb.save(bom_path)
    wb.close()


def write_object_to_excel(bom_path, object_element, sheet_name):

    wb = openpyxl.load_workbook(bom_path)
    type(wb)

    sheet = wb.get_sheet_by_name(sheet_name)
    object_value_list = object_element.values_to_list()
    max_row = sheet.max_row

    if sheet.max_row == 1:
        print("ilosć wierszy to 1, wartość 1 komorki tego wiersza to:")
        print(sheet.cell(row=max_row, column=1).value)
        sheet.cell(row=max_row + 1, column=1).value = 1
        print("wartość dodanej drugiej komórki to: ", sheet.cell(row=max_row+1, column=1).value)
    else:
        last_value = sheet.cell(row=max_row, column=1).value
        print(last_value)

        sheet.cell(row=max_row+1, column=1).value = int(last_value) + 1

    for j in range(len(object_value_list)):
        sheet.cell(row=max_row+1, column=j+2).value = object_value_list[j]

    wb.save(bom_path)
    wb.close()


def write_list_to_excel(bom_path, object_list, sheets_names):
    print("\n\n Otrzymano liste objektów.")
    if len(object_list) > 0:
        if object_list[0].tch.upper() == "C":
            sheet_name = sheets_names[0]
        elif object_list[0].tch.upper() == "F":
            sheet_name = sheets_names[1]
        elif object_list[0].tch.upper() == "S":
            sheet_name = sheets_names[2]
        elif "DRUK" in object_list[0].tch.upper() and "3D" in object_list[0].tch.upper():
            sheet_name = sheets_names[3]
        elif object_list[0].tch.upper() == "Z":
            sheet_name = sheets_names[4]
        else:
            sheet_name = sheets_names[5]

        print("liste obiektów wpisujemy do arkusza: ", sheet_name)

        for element in object_list:
            # zamiast nowej funkcji wystarczy przenieść jej zawartosć tutaj tak zeby nie trzeba bylo za kazdym razem otwierac excela tylko za jednym razem wysłać
            # przyspieszy to dzialanie programu,
            # na upartego do funkcji write_list_to_excel dac argument otwarty typu lista i wpisać do niego wszystkie listy od razu i w tej pętli te listy wpisać do excela
            write_object_to_excel(bom_path, element, sheet_name)

    else:
        print("Otrzymana lista jest pusta.")
        return 0
    return 0


def consolidation_and_segregation(bom_path):
    sheets_name = ["Blachy", "Frezowanie toczenie", "Spawane", "DRUK 3D", "Z-normalia", "Zakupowe-reszta"]
    title_location = {"part_number": 0, "qty_total": 0, "description": 0, "description2": 0, "rysunek": 0, "tch": 0, "producent": 0,
                      "kod_producenta": 0, "kolor": 0}
    title_names = ["Nr"]
    for element in title_location.keys():
        if element != "tch" and element != "rysunek":
            title_names.append(element)

    for name in sheets_name:
        worksheet_maker(bom_path, name, title_names)

    wb = openpyxl.load_workbook(bom_path)
    type(wb)
    arkusze = wb.sheetnames
    sheet = wb[arkusze[0]]

    for i in range(1, sheet.max_column + 1):
        value = sheet.cell(row=1, column=i).value

        if "PART" in value.upper() and "NUMBER" in value.upper():
            title_location["part_number"] = i
        if "QTY" in value.upper() and "TOTAL" in value.upper():
            title_location["qty_total"] = i
        if "DESCRIPTION" in value.upper() and "2" not in value.upper():
            title_location["description"] = i
        if "DESCRIPTION" in value.upper() and "2" in value.upper():
            title_location["description2"] = i
        if "TCH" in value.upper() and "1" in value.upper():
            title_location["tch"] = i
        if "PRODUCENT" in value.upper() and "KOD" not in value.upper():
            title_location["producent"] = i
        if "KOD" in value.upper() and "PRODUCENTA" in value.upper():
            title_location["kod_producenta"] = i
        if "KOLOR" in value.upper():
            title_location["kolor"] = i
        if "RYSUNEK" in value.upper():
            title_location["rysunek"] = i

    print("\n Nagłówki i numery odpowiadających im kolumn w pierwszym arkuszu: ")
    print(title_location)
    for value in title_location.values():
        if value == 0:
            print("Nie udało sie odnaleść lokalizacji jednego z nagłówków w BOM!!!")

    object_list = []
    print(sheet.max_row)
    for row_number in range(2, sheet.max_row + 1):
        print(row_number)
#        input("nacisnij klawisz aby przejsc dalej.")
        if "ZŁOŻENIOWY" not in sheet.cell(row=row_number, column=title_location["rysunek"]).value.upper():
            part_number = sheet.cell(row=row_number, column=title_location["part_number"]).value.lstrip()
            print(part_number)
            qty_total = int(sheet.cell(row=row_number, column=title_location["qty_total"]).value)
            description = sheet.cell(row=row_number, column=title_location["description"]).value
            description2 = sheet.cell(row=row_number, column=title_location["description2"]).value
            tch = sheet.cell(row=row_number, column=title_location["tch"]).value
            producent = sheet.cell(row=row_number, column=title_location["producent"]).value
            kod_producenta = sheet.cell(row=row_number, column=title_location["kod_producenta"]).value
            kolor = sheet.cell(row=row_number, column=title_location["kolor"]).value
            a = Part(part_number, qty_total, description, description2, tch, producent, kod_producenta, kolor)

# sprawdzenie czy obiekt o takim samym PART_NUMBER jus istnieje w tablicy
# jak istnieje stosujemy metode add_to_qty_total ==> dodaje do istniejącego obiektu qty total obecnego (konsolidacja)
# jak nie istnieje to dodajemy element do listy obiektów
        if len(object_list) >= 1:
            status = False
            for element in object_list:
                if element.part_number == part_number:
                    print("Obiekt o takim numerze części już zostal dodany")
                    element.add_to_qty_total(qty_total)
                    status = True
                    break
                else:
                    status = False
            if not status:
                print("dodajemy objekt do listy")
                a.print_values()
                object_list.append(a)

        else:
            print("dodajemy pierwszy obiekt do listy")
            object_list.append(a)

    wb.save(bom_path)
    wb.close()

    object_list_c = []
    object_list_f = []
    object_list_s = []
    object_list_druk3d = []
    object_list_z_normalia = []
    object_list_zakupowe_reszta = []
    print("\n\n segregujemy obiekty zgodnie z tch:\n\n")
    for i in object_list:
        i.print_values()
        if i.tch.upper() == "C":
            print("wrzucamy do C")
            object_list_c.append(i)
        elif i.tch.upper() == "F":
            print("wrzucamy do F")
            object_list_f.append(i)
        elif i.tch.upper() == "S":
            print("wrzucamy do S")
            object_list_s.append(i)
        elif i.tch.upper() == "DRUK 3D":
            print("wrzucamy do druk 3d")
            object_list_druk3d.append(i)
        elif i.tch.upper() == "Z":
            print("wrzucamy do Z")
            object_list_z_normalia.append(i)
        else:
            print("wrzucamy do zakupowe reszta")
            object_list_zakupowe_reszta.append(i)

    write_list_to_excel(bom_path, object_list_c, sheets_name)
    write_list_to_excel(bom_path, object_list_f, sheets_name)
    write_list_to_excel(bom_path, object_list_s, sheets_name)
    write_list_to_excel(bom_path, object_list_druk3d, sheets_name)
    write_list_to_excel(bom_path, object_list_z_normalia, sheets_name)
    write_list_to_excel(bom_path, object_list_zakupowe_reszta, sheets_name)
